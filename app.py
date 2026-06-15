import json
import os
import re
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.request
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
import xml.etree.ElementTree as ET

from flask import Flask, Response, jsonify, render_template, request
from openpyxl import load_workbook
from werkzeug.exceptions import HTTPException


BASE_DIR = Path(__file__).parent.resolve()
SCRIPTS_DIR = BASE_DIR / "scripts"
PROMPTS_DIR = BASE_DIR / "backend_prompts"
GENERATED_SCRIPTS_DIR = Path(
    os.environ.get("GENERATED_SCRIPTS_DIR")
    or (r"C:\Users\Coschool\downloads\scripts" if os.name == "nt" else "/tmp/manim_scripts")
).expanduser()
RENDER_LAUNCH_LOG_DIR = BASE_DIR / "logs" / "render_launcher"
DEFAULT_RENDER_QUALITY = "m"
MAX_REGENERATIONS = 3
DEFAULT_CLAUDE_MODEL = "claude-opus-4-6"
DEFAULT_GEMMA_MODEL = "gemma-4-26b-a4b-it"
GEMMA_MODELS = ["gemma-4-26b-a4b-it", "gemma-4-31b-it"]
MAX_LOS_PER_GROUPED_SCRIPT = 3
PHASE2_READY_STATUSES = {"Approved", "Needs Minor Revision"}

PROMPT_FILES = {
    "script_generation": "conceptual_script_generation.txt",
    "learning_design": "learning_design.txt",
    "lo_grouping": "lo_grouping.txt",
    "grouping_validation": "grouping_validation.txt",
    "validation": "validation_prompt.txt",
    "animation_phase": "animation_phase_master_prompt.txt",
}

PROMPT_META = {
    "lo_grouping": {
        "title": "LO Grouping Prompt",
        "stage": "Step 1",
        "purpose": "Groups selected Learning Outcomes into 2-minute conceptual script clusters.",
    },
    "grouping_validation": {
        "title": "Grouping Validation Prompt",
        "stage": "Step 1 QA",
        "purpose": "Validates LO grouping against completeness, no-duplication, and no-cross-CC rules.",
    },
    "script_generation": {
        "title": "Generation Prompt Conceptual Script",
        "stage": "Step 2",
        "purpose": "Generates voice-over-ready conceptual scripts from grouped Learning Outcomes.",
    },
    "learning_design": {
        "title": "Learning Design Final",
        "stage": "Blueprint",
        "purpose": "Defines the instructional framework used inside script generation.",
    },
    "validation": {
        "title": "Validation Prompt v2 Updated",
        "stage": "Step 3",
        "purpose": "Validates scripts and drives capped regeneration until approval.",
    },
    "animation_phase": {
        "title": "Animation Phase Master Prompt",
        "stage": "Phase 2",
        "purpose": "Turns approved transcripts into storyboard, position table, and Manim Python code.",
    },
}

app = Flask(__name__)
JOB_EXECUTOR = ThreadPoolExecutor(max_workers=4)
JOB_LOCK = threading.Lock()
JOBS = {}
JOB_TTL_SECONDS = 60 * 60 * 12


@app.errorhandler(Exception)
def json_error_handler(exc):
    if isinstance(exc, HTTPException):
        original = getattr(exc, "original_exception", None)
        if original:
            return jsonify({"error": str(original) or original.__class__.__name__}), exc.code
        return jsonify({"error": exc.description or exc.name}), exc.code
    return jsonify({"error": str(exc) or exc.__class__.__name__}), 500


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Claude-Key, X-Gemma-Key"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


def load_prompt(name):
    path = PROMPTS_DIR / PROMPT_FILES[name]
    return path.read_text(encoding="utf-8")


def fill_prompt(template, **values):
    output = template
    for key, value in values.items():
        output = output.replace("{{" + key + "}}", str(value))
        output = output.replace("{{" + key.upper() + "}}", str(value))
    return output


def get_claude_api_key():
    api_key = request.headers.get("X-Claude-Key", "").strip()
    if not api_key:
        payload = request.get_json(silent=True) or {}
        api_key = str(payload.get("claude_api_key", "")).strip()
    if not api_key:
        api_key = str(request.form.get("claude_api_key", "")).strip()
    if not api_key:
        raise ValueError("Missing Claude API key. Save your Claude key before running Phase 2.")
    return api_key


def get_gemma_api_key():
    api_key = request.headers.get("X-Gemma-Key", "").strip()
    if not api_key:
        payload = request.get_json(silent=True) or {}
        api_key = str(payload.get("gemma_api_key", "")).strip()
    if not api_key:
        api_key = str(request.form.get("gemma_api_key", "")).strip()
    if not api_key:
        raise ValueError("Missing Gemma API key. Save your Gemma key before running.")
    return api_key


def claude_api_key_from_payload(payload):
    api_key = str((payload or {}).get("claude_api_key", "")).strip()
    if not api_key:
        raise ValueError("Missing Claude API key. Save your Claude key before running Phase 2.")
    return api_key


def gemma_api_key_from_payload(payload):
    api_key = str((payload or {}).get("gemma_api_key", "")).strip()
    if not api_key:
        raise ValueError("Missing Gemma API key. Save your Gemma key before running.")
    return api_key


def model_provider_from_payload(payload):
    provider = str((payload or {}).get("modelProvider") or "claude").strip().lower()
    return "gemma" if provider == "gemma" else "claude"


def model_name_from_payload(payload):
    provider = model_provider_from_payload(payload)
    if provider == "gemma":
        return str((payload or {}).get("gemmaModel") or DEFAULT_GEMMA_MODEL).strip() or DEFAULT_GEMMA_MODEL
    return str((payload or {}).get("claudeModel") or DEFAULT_CLAUDE_MODEL).strip() or DEFAULT_CLAUDE_MODEL


def payload_with_request_key(payload):
    output = dict(payload or {})
    provider = model_provider_from_payload(output)
    output["modelProvider"] = provider
    if provider == "gemma":
        if not str(output.get("gemma_api_key", "")).strip():
            try:
                output["gemma_api_key"] = get_gemma_api_key()
            except Exception:
                pass
        return output
    if not str(output.get("claude_api_key", "")).strip():
        try:
            output["claude_api_key"] = get_claude_api_key()
        except Exception:
            pass
    return output


def call_claude(prompt, api_key, model=None, max_output_tokens=16000):
    body = json.dumps(
        {
            "model": model or DEFAULT_CLAUDE_MODEL,
            "max_tokens": max_output_tokens,
            "messages": [{"role": "user", "content": prompt}],
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={
            "content-type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ValueError(f"Claude API request failed: {detail}") from exc
    except urllib.error.URLError as exc:
        raise ValueError(f"Claude API request failed: {exc.reason}") from exc

    text_parts = []
    for block in data.get("content", []):
        if block.get("type") == "text":
            text_parts.append(block.get("text", ""))
    output = "\n".join(text_parts).strip()
    if not output:
        raise ValueError("Claude returned an empty response.")
    return output


def call_gemma(prompt, api_key, model=None, max_output_tokens=16000):
    selected_model = str(model or DEFAULT_GEMMA_MODEL).strip()
    selected_model = selected_model.removeprefix("models/")
    body = json.dumps(
        {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"maxOutputTokens": max_output_tokens},
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        f"https://generativelanguage.googleapis.com/v1beta/models/{selected_model}:generateContent?key={api_key}",
        data=body,
        headers={"content-type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ValueError(f"Gemma API request failed: {detail}") from exc
    except urllib.error.URLError as exc:
        raise ValueError(f"Gemma API request failed: {exc.reason}") from exc

    text_parts = []
    for candidate in data.get("candidates", []):
        content = candidate.get("content") or {}
        for part in content.get("parts", []):
            if "text" in part:
                text_parts.append(part.get("text", ""))
    output = "\n".join(text_parts).strip()
    if not output:
        raise ValueError("Gemma returned an empty response.")
    return output


def call_model(prompt, payload, max_output_tokens=16000):
    provider = model_provider_from_payload(payload)
    model = model_name_from_payload(payload)
    if provider == "gemma":
        return call_gemma(prompt, gemma_api_key_from_payload(payload), model, max_output_tokens=max_output_tokens)
    return call_claude(prompt, claude_api_key_from_payload(payload), model, max_output_tokens=max_output_tokens)


def normalise_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def find_column(headers, candidates):
    normalised = {normalise_header(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        key = normalise_header(candidate)
        if key in normalised:
            return normalised[key]
    return None


def find_columns(headers, candidates):
    normalised = {normalise_header(header): index for index, header in enumerate(headers)}
    matches = []
    for candidate in candidates:
        key = normalise_header(candidate)
        if key in normalised and normalised[key] not in matches:
            matches.append(normalised[key])
    return matches


def find_header_row(rows):
    best_index = None
    best_score = -1
    for index, row in enumerate(rows[:25]):
        headers = [str(cell or "").strip() for cell in row]
        score = 0
        if find_column(headers, GRADE_COLUMNS) is not None:
            score += 1
        if find_column(headers, CHAPTER_COLUMNS) is not None:
            score += 3
        if find_column(headers, SUBTOPIC_COLUMNS) is not None:
            score += 1
        if find_column(headers, LO_COLUMNS) is not None:
            score += 3
        if find_column(headers, SUBJECT_COLUMNS) is not None:
            score += 1
        if score > best_score:
            best_index = index
            best_score = score
    if best_score >= 6:
        return best_index
    return None


def row_value(row, index):
    if isinstance(index, (list, tuple)):
        for item in index:
            value = row_value(row, item)
            if value:
                return value
        return ""
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


GRADE_COLUMNS = ["gradeName", "Grade Name", "Grade", "Class", "Class Name", "Grade Level", "Class Level"]
SUBJECT_COLUMNS = ["subjectName", "Subject Name", "Subject"]
CHAPTER_COLUMNS = [
    "topicName",
    "Topic Name",
    "Chapter Name",
    "Chapter Names",
    "Chapter Title",
    "Chapter",
    "Topic",
    "Unit Name",
    "Unit",
    "Lesson",
    "Lesson Name",
    "Session",
    "Session Name",
]
SUBTOPIC_COLUMNS = [
    "subTopicName",
    "Subtopic Name",
    "Sub Topic",
    "Sub-Topic",
    "Subtopic",
    "Subtopic Title",
    "KnowledgeCellName",
    "Knowledge Cell Name",
    "Concept Name",
    "Concept",
]
CC_COLUMNS = ["KnowledgeCellId", "Knowledge Cell Id", "CC", "C.C.", "Competency Code", "Concept Code", "Content Code", "Chapter Code"]
CC_NAME_COLUMNS = ["KnowledgeCellName", "Knowledge Cell Name", "CC Name", "Competency Name", "Competency", "Concept Name", "Concept", "Content"]
LO_COLUMNS = [
    "LearningOutcomeName",
    "Learning Outcome Name",
    "Learning Outcome",
    "Learning Outcomes",
    "LO",
    "LO Text",
    "Outcome",
    "Learning Objective",
    "Learning Objectives",
    "LearningOutcomes",
    "LearningOutcomesText",
]


def local_xml_name(tag):
    return tag.rsplit("}", 1)[-1]


def cell_column_index(cell_ref):
    letters = "".join(char for char in str(cell_ref or "") if char.isalpha())
    if not letters:
        return None
    index = 0
    for char in letters.upper():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def load_xlsx_shared_strings(archive):
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    shared_strings = []
    with archive.open("xl/sharedStrings.xml") as shared_xml:
        for event, elem in ET.iterparse(shared_xml, events=("end",)):
            if local_xml_name(elem.tag) != "si":
                continue
            text = "".join(node.text or "" for node in elem.iter() if local_xml_name(node.tag) == "t")
            shared_strings.append(text)
            elem.clear()
    return shared_strings


def read_xlsx_cell(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter() if local_xml_name(node.tag) == "t").strip()

    raw_value = ""
    for child in cell:
        if local_xml_name(child.tag) == "v":
            raw_value = child.text or ""
            break
    if not raw_value:
        return ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)].strip()
        except (IndexError, TypeError, ValueError):
            return raw_value.strip()
    if cell_type == "b":
        return "TRUE" if raw_value == "1" else "FALSE"
    return raw_value.strip()


def iter_xlsx_sheet_rows(workbook_path, sheet_path, shared_strings):
    with zipfile.ZipFile(workbook_path) as archive:
        with archive.open(sheet_path) as sheet_xml:
            fallback_row_number = 0
            for event, elem in ET.iterparse(sheet_xml, events=("end",)):
                if local_xml_name(elem.tag) != "row":
                    continue
                fallback_row_number += 1
                row_number = int(elem.attrib.get("r") or fallback_row_number)
                row_values = []
                for cell in elem:
                    if local_xml_name(cell.tag) != "c":
                        continue
                    column_index = cell_column_index(cell.attrib.get("r"))
                    if column_index is None:
                        column_index = len(row_values)
                    if column_index >= len(row_values):
                        row_values.extend([""] * (column_index - len(row_values) + 1))
                    row_values[column_index] = read_xlsx_cell(cell, shared_strings)
                yield row_number, tuple(row_values)
                elem.clear()


def preview_sheet_rows(row_iterator_factory, limit=25):
    preview = []
    for row_number, row in row_iterator_factory():
        preview.append((row_number, row))
        if len(preview) >= limit:
            break
    return preview


def parse_sheet_rows(row_iterator_factory, preview_pairs=None):
    if preview_pairs is None:
        preview_pairs = preview_sheet_rows(row_iterator_factory)
    preview_rows = [row for row_number, row in preview_pairs]
    if not preview_rows:
        return []

    header_position = find_header_row(preview_rows)
    if header_position is None:
        preview_headers = [str(cell or "").strip() for cell in preview_rows[0] if str(cell or "").strip()]
        raise ValueError(
            "Could not find the header row. Expected columns like Chapter/Topic and Learning Outcome. "
            + "Optional columns include Grade/Class, Subject, and Subtopic/Sub Topic. "
            + "First row detected: "
            + (", ".join(preview_headers) if preview_headers else "blank")
        )

    header_row_number, header_row = preview_pairs[header_position]
    headers = [str(cell or "").strip() for cell in header_row]
    columns = {
        "grade": find_columns(headers, GRADE_COLUMNS),
        "subject": find_columns(headers, SUBJECT_COLUMNS),
        "chapter": find_columns(headers, CHAPTER_COLUMNS),
        "subtopic": find_columns(headers, SUBTOPIC_COLUMNS),
        "cc": find_columns(headers, CC_COLUMNS),
        "cc_name": find_columns(headers, CC_NAME_COLUMNS),
        "lo": find_columns(headers, LO_COLUMNS),
    }

    required = ["chapter", "lo"]
    missing = [name for name in required if not columns[name]]
    if missing:
        detected = [header for header in headers if header]
        raise ValueError(
            "Missing required column(s): "
            + ", ".join(missing)
            + ". Detected headers: "
            + (", ".join(detected) if detected else "none")
        )

    parsed = []
    last_values = {
        "grade": "All Grades",
        "subject": "Mathematics",
        "chapter": "",
        "subtopic": "",
        "cc": "",
        "cc_name": "",
    }
    seen_learning_outcomes = set()
    relevant_column_indexes = sorted(
        {
            index
            for indexes in columns.values()
            for index in indexes
            if index is not None
        }
    )
    blank_relevant_rows = 0
    for row_number, row in row_iterator_factory():
        if row_number <= header_row_number:
            continue
        row_has_relevant_value = any(row_value(row, index) for index in relevant_column_indexes)
        if not row_has_relevant_value:
            blank_relevant_rows += 1
            if parsed and blank_relevant_rows >= 100:
                break
            if not parsed and blank_relevant_rows >= 1000:
                break
            continue
        blank_relevant_rows = 0
        for field in last_values:
            value = row_value(row, columns.get(field))
            if value:
                last_values[field] = value
        if not last_values["chapter"]:
            continue
        lo_text = row_value(row, columns["lo"])
        if not lo_text:
            continue
        item_key = (
            last_values["grade"],
            last_values["subject"],
            last_values["chapter"],
            last_values["subtopic"],
            lo_text,
        )
        if item_key in seen_learning_outcomes:
            continue
        seen_learning_outcomes.add(item_key)
        parsed.append(
            {
                "row": row_number,
                "grade": last_values["grade"] or "All Grades",
                "subject": last_values["subject"] or "Mathematics",
                "chapter": last_values["chapter"],
                "subtopic": last_values["subtopic"],
                "cc": last_values["cc"],
                "cc_name": last_values["cc_name"],
                "lo": lo_text,
                "source": "learning_outcome",
            }
        )
    if not parsed:
        raise ValueError("No Learning Outcome rows found below the detected header row.")
    return parsed


def xlsx_sheet_paths(workbook_path):
    with zipfile.ZipFile(workbook_path) as archive:
        return sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            ]
        )


def parse_xlsx_fast(workbook_path):
    with zipfile.ZipFile(workbook_path) as archive:
        shared_strings = load_xlsx_shared_strings(archive)
    sheet_paths = xlsx_sheet_paths(workbook_path)
    if not sheet_paths:
        raise ValueError("No worksheets found in the uploaded Excel file.")

    first_preview = None
    first_sheet_path = sheet_paths[0]
    for sheet_path in sheet_paths:
        row_factory = lambda path=sheet_path: iter_xlsx_sheet_rows(workbook_path, path, shared_strings)
        preview_pairs = preview_sheet_rows(row_factory)
        if first_preview is None:
            first_preview = preview_pairs
        if find_header_row([row for row_number, row in preview_pairs]) is not None:
            return parse_sheet_rows(row_factory, preview_pairs)

    first_factory = lambda path=first_sheet_path: iter_xlsx_sheet_rows(workbook_path, path, shared_strings)
    return parse_sheet_rows(first_factory, first_preview or [])


def iter_openpyxl_sheet_rows(sheet):
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        yield row_number, row


def parse_excel_openpyxl(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for candidate_sheet in workbook.worksheets:
            row_factory = lambda sheet=candidate_sheet: iter_openpyxl_sheet_rows(sheet)
            preview_pairs = preview_sheet_rows(row_factory)
            if find_header_row([row for row_number, row in preview_pairs]) is not None:
                return parse_sheet_rows(row_factory, preview_pairs)
        active_factory = lambda: iter_openpyxl_sheet_rows(workbook.active)
        return parse_sheet_rows(active_factory)
    finally:
        workbook.close()


def parse_excel_path(workbook_path):
    workbook_path = Path(workbook_path)
    try:
        return parse_xlsx_fast(workbook_path)
    except (ET.ParseError, KeyError, OSError, zipfile.BadZipFile):
        return parse_excel_openpyxl(workbook_path)


def parse_excel(file_storage):
    if isinstance(file_storage, (str, Path)):
        return parse_excel_path(file_storage)

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = Path(temp_file.name)
            if hasattr(file_storage, "save"):
                file_storage.save(temp_file)
            else:
                temp_file.write(file_storage.read())
        return parse_excel_path(temp_path)
    finally:
        if temp_path:
            temp_path.unlink(missing_ok=True)


def parse_transcript_file(file_storage):
    filename = str(getattr(file_storage, "filename", "") or "").strip()
    suffix = Path(filename).suffix.lower()
    data = file_storage.read()
    if not data:
        raise ValueError("Uploaded transcript file is empty.")

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise ValueError("PDF transcript support is not installed. Install pypdf and retry.") from exc
        reader = PdfReader(BytesIO(data))
        pages = []
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages.append(text.strip())
        transcript = "\n\n".join(pages).strip()
        if not transcript:
            raise ValueError("No readable transcript text was found in the PDF.")
        return transcript

    if suffix not in {".txt", ".text", ".md", ".markdown", ""}:
        raise ValueError("Upload a transcript as .txt, .md, or .pdf.")

    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return data.decode(encoding).strip()
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode the transcript text file.")


def filters_for(rows):
    def unique(key, items):
        return sorted({row[key] for row in items if row.get(key)})

    grades = unique("grade", rows)
    all_chapters = unique("chapter", rows)
    all_subtopics = unique("subtopic", rows)
    chapters_by_grade = {}
    subtopics_by_chapter = {}
    for grade in grades:
        grade_rows = [row for row in rows if row["grade"] == grade]
        chapters_by_grade[grade] = unique("chapter", grade_rows)
        for chapter in chapters_by_grade[grade]:
            chapter_key = f"{grade}||{chapter}"
            subtopics_by_chapter[chapter_key] = unique(
                "subtopic", [row for row in grade_rows if row["chapter"] == chapter]
            )
    return {
        "grades": grades,
        "allChapters": all_chapters,
        "allSubtopics": all_subtopics,
        "chaptersByGrade": chapters_by_grade,
        "subtopicsByChapter": subtopics_by_chapter,
    }


def format_ccs_and_los(outcomes):
    grouped = {}
    for index, outcome in enumerate(outcomes, start=1):
        source_key = outcome.get("cc") or outcome.get("subtopic") or outcome.get("chapter") or "Visible Learning Outcomes"
        source_name = outcome.get("cc_name") or outcome.get("subtopic") or "Learning Outcomes"
        grouped.setdefault((source_key, source_name), []).append((index, outcome))

    blocks = []
    for group_index, ((source_key, source_name), items) in enumerate(grouped.items(), start=1):
        lines = [f"Source Group {group_index}: {source_key} {source_name}".strip()]
        lines.extend(
            f"LO {index}. {outcome.get('lo', '')}"
            for index, outcome in items
            if outcome.get("lo")
        )
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks)


def extract_verdict(report):
    verdict_match = re.search(r"#?\s*VERDICT\s*(.*)$", report, flags=re.IGNORECASE | re.DOTALL)
    verdict_text = verdict_match.group(1) if verdict_match else report[-1000:]
    verdict_text = verdict_text.lower()
    if "approved" in verdict_text and "ready" in verdict_text:
        return "Approved"
    if "needs minor revision" in verdict_text or "minor revision" in verdict_text:
        return "Needs Minor Revision"
    if "rejected" in verdict_text or "regenerate" in verdict_text:
        return "Rejected"
    if "approved" in verdict_text:
        return "Approved"
    return "Needs Review"


def is_phase2_ready_status(status):
    return str(status or "").strip() in PHASE2_READY_STATUSES


def script_prompt(grade, chapter, subtopic, outcomes_text):
    generation = fill_prompt(
        load_prompt("script_generation"),
        Grade=grade,
        ChapterName=chapter,
        SubtopicName=subtopic,
        LearningOutcomes=outcomes_text,
    )
    learning_design = load_prompt("learning_design")
    return (
        "Use the following Learning Design as the instructional blueprint.\n\n"
        f"{learning_design}\n\n"
        "Now generate the conceptual script using this generation prompt.\n\n"
        f"{generation}"
    )


def validation_prompt(grade, chapter, subtopic, outcomes_text, script, textbook_reference=""):
    textbook_reference = textbook_reference or (
        "No textbook PDF has been attached in Phase 1. Validate against the selected Learning Outcomes, "
        "the Learning Design, and mathematical correctness. Do not invent page references."
    )
    return fill_prompt(
        load_prompt("validation"),
        GRADE=grade,
        TOPIC=chapter,
        SUBTOPIC=subtopic,
        LEARNING_OUTCOMES=outcomes_text,
        SCRIPT_TO_REVIEW=script,
        ATTACHED_TEXTBOOK_PDF=textbook_reference,
    )


def animation_metadata(payload):
    title = str(payload.get("title") or payload.get("chapter") or "Concept Animation").strip()
    return {
        "title": title,
        "subject": str(payload.get("subject") or "Mathematics").strip(),
        "grade": str(payload.get("grade") or "").strip(),
        "duration": str(payload.get("duration") or "90 seconds").strip(),
        "content_type": str(payload.get("contentType") or "Concept Video").strip(),
    }


def animation_prompt(payload, mode):
    transcript = str(payload.get("transcript") or "").strip()
    if not transcript:
        raise ValueError("Paste a transcript or use a Phase 2 ready script first.")

    meta = animation_metadata(payload)
    master_prompt = load_prompt("animation_phase")
    stage_instruction = {
        "storyboard": (
            "Run only STEP 0, STEP 1, STEP 2, STEP 3a, STEP 3b, and the pre-code/audit planning items "
            "needed for the storyboard. Do not output Python code."
        ),
        "code": (
            "Using the transcript and storyboard below, output STEP 4 only: raw production-ready Python code. "
            "Do not wrap the code in markdown fences. Do not include prose before or after the Python."
        ),
        "package": (
            "Run the full pipeline in the final output order: STEP 0, STEP 1, STEP 2, STEP 3a, STEP 3b, "
            "STEP 4, STEP 5, and STEP 6."
        ),
    }[mode]

    storyboard = str(payload.get("storyboard") or "").strip()
    storyboard_block = f"\n\nEXISTING STORYBOARD PACKAGE:\n{storyboard}\n" if storyboard else ""
    return (
        f"{master_prompt}\n\n"
        "USER INPUT FOR THIS RUN:\n"
        f'INPUT:\n"""\n{transcript}\n"""\n\n'
        "METADATA:\n"
        f"Title: {meta['title']}\n"
        f"Subject: {meta['subject']}\n"
        f"Grade: {meta['grade']}\n"
        f"Duration: {meta['duration']}\n"
        f"Content Type: {meta['content_type']}\n"
        f"{storyboard_block}\n"
        "EXECUTION MODE:\n"
        f"{stage_instruction}"
    )


def extract_python_code(text):
    fenced = re.search(r"```(?:python)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    return fenced.group(1).strip() if fenced else text.strip()


def slugify_filename(value, fallback="animation"):
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return slug[:80] or fallback


def unique_python_filename(title, used_names):
    base = slugify_filename(title, "animation_scene")
    filename = f"{base}.py"
    counter = 2
    while filename in used_names:
        filename = f"{base}_{counter}.py"
        counter += 1
    used_names.add(filename)
    return filename


def unique_python_path(directory, title, used_names=None):
    used_names = used_names if used_names is not None else set()
    base = slugify_filename(title, "animation_scene")
    counter = 1
    while True:
        filename = f"{base}.py" if counter == 1 else f"{base}_{counter}.py"
        path = directory / filename
        if filename not in used_names and not path.exists():
            used_names.add(filename)
            return path
        counter += 1


def save_generated_python_script(title, code, used_names=None):
    clean_code = extract_python_code(code)
    if not clean_code.strip():
        raise ValueError("Generated Python code was empty.")
    GENERATED_SCRIPTS_DIR.mkdir(parents=True, exist_ok=True)
    path = unique_python_path(GENERATED_SCRIPTS_DIR, title, used_names)
    path.write_text(clean_code.rstrip() + "\n", encoding="utf-8")
    return {
        "filename": path.name,
        "path": str(path),
    }


def launch_start_renderer(saved_files, quality=DEFAULT_RENDER_QUALITY):
    paths = []
    for item in saved_files or []:
        raw_path = item.get("path") if isinstance(item, dict) else item
        if not raw_path:
            continue
        path = Path(raw_path)
        if path.exists() and path.suffix.lower() == ".py":
            paths.append(path)
    if not paths:
        return {"started": False, "error": "No saved Python scripts to render."}

    RENDER_LAUNCH_LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = RENDER_LAUNCH_LOG_DIR / f"start_{int(time.time())}_{uuid.uuid4().hex[:8]}.log"
    filenames = ",".join(path.name for path in paths)
    command = [
        sys.executable,
        str(BASE_DIR / "start.py"),
        "--folder",
        str(GENERATED_SCRIPTS_DIR),
        "--files",
        filenames,
        "--quality",
        quality or DEFAULT_RENDER_QUALITY,
        "--yes",
        "--no-open",
    ]
    env = {**os.environ, "MANIM_SCRIPTS_DIR": str(GENERATED_SCRIPTS_DIR)}
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        with log_path.open("w", encoding="utf-8") as log_file:
            process = subprocess.Popen(
                command,
                cwd=str(BASE_DIR),
                stdout=log_file,
                stderr=subprocess.STDOUT,
                env=env,
                creationflags=creationflags,
            )
    except Exception as exc:
        return {
            "started": False,
            "files": [path.name for path in paths],
            "log": str(log_path),
            "error": str(exc) or exc.__class__.__name__,
        }
    return {
        "started": True,
        "pid": process.pid,
        "files": [path.name for path in paths],
        "log": str(log_path),
        "quality": quality or DEFAULT_RENDER_QUALITY,
    }


def animation_title_for_outcome(outcome):
    return " - ".join(
        str(part)
        for part in [outcome.get("chapter"), outcome.get("subtopic"), outcome.get("cc")]
        if part
    ) or "Concept Animation"


def script_animation_title(index, group):
    return f"Script {index} - {animation_title_for_outcome(group)}"


def bulk_groups_from_outcomes(outcomes):
    grouped = {}
    for outcome in outcomes:
        key = (
            outcome.get("grade") or "",
            outcome.get("subject") or "Mathematics",
            outcome.get("chapter") or "",
            outcome.get("subtopic") or "",
            outcome.get("cc") or "",
            outcome.get("cc_name") or "",
        )
        if key not in grouped:
            grouped[key] = {
                "grade": key[0],
                "subject": key[1],
                "chapter": key[2],
                "subtopic": key[3],
                "cc": key[4],
                "cc_name": key[5],
                "rows": [],
                "outcomes": [],
            }
        grouped[key]["rows"].append(outcome.get("row"))
        grouped[key]["outcomes"].append(outcome)
    return list(grouped.values())


def bulk_learning_outcomes_text(outcomes):
    return "\n".join(
        f"LO {index}. {outcome.get('lo', '')}"
        for index, outcome in enumerate(outcomes, start=1)
        if outcome.get("lo")
    )


def grouping_validation_prompt(grade, chapter, ccs_and_los, grouping_table):
    return fill_prompt(
        load_prompt("grouping_validation"),
        Grade=grade,
        ChapterName=chapter,
        CCsAndLOs=ccs_and_los,
        GroupingTable=grouping_table,
    )


def grouping_validation_status(report):
    verdict = str(report or "").upper()
    if "REJECTED" in verdict:
        return "Rejected"
    if "NEEDS MINOR REVISION" in verdict:
        return "Needs Minor Revision"
    if "APPROVED" in verdict:
        return "Approved"
    return "Needs Review"


def grouped_script_blocks(grouping_text):
    text = str(grouping_text or "").strip()
    if not text:
        return []

    def script_key(title):
        match = re.search(r"\bScript\s+(\d+)\b", str(title or ""), re.I)
        if match:
            return f"script-{int(match.group(1)):04d}"
        return re.sub(r"[^a-z0-9]+", "", str(title or "").lower()) or str(len(text))

    def ordered_unique(blocks):
        by_key = {}
        order = []
        for block in blocks:
            key = script_key(block.get("title"))
            if key not in by_key:
                order.append(key)
                by_key[key] = block
            elif len(block.get("learningOutcomes", "")) > len(by_key[key].get("learningOutcomes", "")):
                by_key[key] = block
        return [by_key[key] for key in order]

    def script_number(block):
        match = re.search(r"\bScript\s+(\d+)\b", str(block.get("title", "")), re.I)
        return int(match.group(1)) if match else 10**9

    marker_blocks = re.findall(
        r"(?is)START_SCRIPT_GROUP\s*(.*?)(?=START_SCRIPT_GROUP|\Z)",
        text,
    )
    parsed_marker_blocks = []
    if marker_blocks:
        for index, block in enumerate(marker_blocks, start=1):
            block = re.sub(r"(?is)\s*END_SCRIPT_GROUP.*$", "", block).strip()
            title_match = re.search(r"(?im)^\s*TITLE:\s*(.+?)\s*$", block)
            title = title_match.group(1).strip() if title_match else f"Script {index}"
            cc_match = re.search(r"(?im)^\s*CC:\s*(.+?)\s*$", block)
            los_match = re.search(r"(?ims)^\s*LOS:\s*(.+)$", block)
            parts = [title]
            if cc_match:
                parts.append(cc_match.group(1).strip())
            if los_match:
                parts.append(los_match.group(1).strip())
            learning_outcomes = "\n".join(part for part in parts if part).strip()
            if learning_outcomes:
                parsed_marker_blocks.append({"title": title, "learningOutcomes": learning_outcomes})

    summary_text = re.split(r"(?im)^\s*GROUPING SCRIPT BLOCKS\s*$", text, maxsplit=1)[0]
    matches = list(
        re.finditer(
            r"(?im)(?:^|\|)\s*(?:#+\s*)?(?:\*\*)?(Script\s+\d+(?:\s+CC\s+[A-Za-z0-9._-]+)?)(?:\*\*)?",
            summary_text,
        )
    )
    if not matches:
        return parsed_marker_blocks or [{"title": "Script 1", "learningOutcomes": text}]

    parsed_summary_blocks = []
    for index, match in enumerate(matches):
        start = summary_text.rfind("\n", 0, match.start()) + 1
        end = matches[index + 1].start() if index + 1 < len(matches) else len(summary_text)
        block = summary_text[start:end].strip()
        title = match.group(1).strip()
        if " CC " not in title:
            cc_match = re.search(r"\bCC\s+[A-Za-z0-9._-]+", block)
            if cc_match:
                title = f"{title} {cc_match.group(0)}"
        if block:
            parsed_summary_blocks.append({"title": title, "learningOutcomes": block})

    combined = ordered_unique(parsed_marker_blocks + parsed_summary_blocks)
    return sorted(combined, key=script_number)


def source_exact_grouped_scripts(groups, outcomes):
    source_by_number = {
        index: outcome
        for index, outcome in enumerate(outcomes, start=1)
        if str(outcome.get("lo") or "").strip()
    }
    exact_groups = []
    assigned_numbers = set()

    def clean_script_title(title, fallback_index):
        match = re.search(r"\bScript\s+(\d+)\b", str(title or ""), re.I)
        if match:
            return f"Script {int(match.group(1))}"
        return f"Script {fallback_index}"

    pending_exact_groups = []
    for group_index, group in enumerate(groups, start=1):
        text = str(group.get("learningOutcomes") or "")
        numbers = []
        for match in re.finditer(r"\bLO\s*0*(\d+)\b", text, re.I):
            number = int(match.group(1))
            if number in source_by_number and number not in numbers and number not in assigned_numbers:
                numbers.append(number)

        if numbers:
            matched = [source_by_number[number] for number in numbers]
            assigned_numbers.update(numbers)
            title = clean_script_title(group.get("title"), group_index)
            lo_lines = [
                f"LO {number}. {source_by_number[number].get('lo', '')}"
                for number in numbers
            ]
            pending_exact_groups.append(
                {
                    **group,
                    "title": title,
                    "learningOutcomes": "\n".join(lo_lines).strip(),
                    "sourceRows": [outcome.get("row") for outcome in matched],
                    "loNumbers": numbers,
                    "cc": "",
                    "cc_name": "",
                    "subtopic": matched[0].get("subtopic") if matched else group.get("subtopic", ""),
                }
            )
        else:
            pending_exact_groups.append({**group, "title": clean_script_title(group.get("title"), group_index), "cc": "", "cc_name": ""})

    missing_numbers = [number for number in source_by_number if number not in assigned_numbers]
    if missing_numbers:
        pending_exact_groups.append(
            {
                "title": f"Script {len(pending_exact_groups) + 1}",
                "learningOutcomes": "\n".join(
                    f"LO {number}. {source_by_number[number].get('lo', '')}"
                    for number in missing_numbers
                ),
                "sourceRows": [source_by_number[number].get("row") for number in missing_numbers],
                "loNumbers": missing_numbers,
                "cc": "",
                "cc_name": "",
                "subtopic": source_by_number[missing_numbers[0]].get("subtopic") if missing_numbers else "",
            }
        )

    for group in pending_exact_groups:
        numbers = list(group.get("loNumbers") or [])
        if not numbers:
            exact_groups.append(group)
            continue
        for start in range(0, len(numbers), MAX_LOS_PER_GROUPED_SCRIPT):
            chunk = numbers[start:start + MAX_LOS_PER_GROUPED_SCRIPT]
            matched = [source_by_number[number] for number in chunk if number in source_by_number]
            if not matched:
                continue
            exact_groups.append(
                {
                    **group,
                    "title": f"Script {len(exact_groups) + 1}",
                    "learningOutcomes": "\n".join(
                        f"LO {number}. {source_by_number[number].get('lo', '')}"
                        for number in chunk
                        if number in source_by_number
                    ).strip(),
                    "sourceRows": [outcome.get("row") for outcome in matched],
                    "loNumbers": chunk,
                    "cc": "",
                    "cc_name": "",
                    "subtopic": matched[0].get("subtopic") if matched else group.get("subtopic", ""),
                }
            )

    for index, group in enumerate(exact_groups, start=1):
        group["title"] = f"Script {index}"
    return exact_groups


@app.get("/")
def index():
    return render_template(
        "index.html",
        max_regenerations=MAX_REGENERATIONS,
        default_claude_model=DEFAULT_CLAUDE_MODEL,
        default_gemma_model=DEFAULT_GEMMA_MODEL,
        gemma_models=GEMMA_MODELS,
    )


@app.get("/health")
def health():
    return jsonify({"status": "ok"})


@app.get("/api")
def api():
    return jsonify(
        {
            "service": "Conceptual Script Generator",
            "status": "ok",
            "phase": "1+2",
            "max_regenerations": MAX_REGENERATIONS,
            "default_claude_model": DEFAULT_CLAUDE_MODEL,
            "default_gemma_model": DEFAULT_GEMMA_MODEL,
            "gemma_models": GEMMA_MODELS,
            "prompts": sorted(PROMPT_FILES),
            "media_storage": "local-only; media/ is intentionally not tracked in Git",
        }
    )


@app.get("/api/prompts")
def prompts_route():
    prompts = []
    for key, meta in PROMPT_META.items():
        text = load_prompt(key)
        prompts.append(
            {
                "key": key,
                "title": meta["title"],
                "stage": meta["stage"],
                "purpose": meta["purpose"],
                "characters": len(text),
            }
        )
    return jsonify({"prompts": prompts})


@app.get("/api/prompts/<name>")
def prompt_detail_route(name):
    if name not in PROMPT_FILES:
        return jsonify({"error": "Unknown prompt."}), 404
    meta = PROMPT_META[name]
    return jsonify(
        {
            "key": name,
            "title": meta["title"],
            "stage": meta["stage"],
            "purpose": meta["purpose"],
            "text": load_prompt(name),
        }
    )


@app.post("/api/validate-key")
@app.post("/api/validate-claude-key")
def validate_claude_key_route():
    try:
        api_key = get_claude_api_key()
        call_claude("Reply with only OK.", api_key, max_output_tokens=16)
    except Exception as exc:
        return jsonify({"valid": False, "error": "Claude API key validation failed: " + str(exc)}), 400
    return jsonify({"valid": True, "message": "Claude API key is valid."})


@app.post("/api/validate-gemma-key")
def validate_gemma_key_route():
    try:
        payload = request.get_json(silent=True) or {}
        api_key = get_gemma_api_key()
        model = str(payload.get("gemmaModel") or request.form.get("gemmaModel") or DEFAULT_GEMMA_MODEL).strip()
        call_gemma("Reply with only OK.", api_key, model, max_output_tokens=16)
    except Exception as exc:
        return jsonify({"valid": False, "error": "Gemma API key validation failed: " + str(exc)}), 400
    return jsonify({"valid": True, "message": "Gemma API key is valid."})


@app.post("/api/parse-excel")
def parse_excel_route():
    if "file" not in request.files:
        return jsonify({"error": "Upload an Excel file first."}), 400
    try:
        rows = parse_excel(request.files["file"])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"rows": rows, "filters": filters_for(rows)})


@app.post("/api/parse-transcript-file")
def parse_transcript_file_route():
    if "file" not in request.files:
        return jsonify({"error": "Upload a transcript file first."}), 400
    try:
        transcript = parse_transcript_file(request.files["file"])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"transcript": transcript})


def group_result(payload):
    outcomes = payload.get("outcomes") or []
    if not outcomes:
        raise ValueError("Select at least one Learning Outcome.")
    grade = payload.get("grade", "")
    chapter = payload.get("chapter", "")
    ccs_and_los = format_ccs_and_los(outcomes)
    prompt = fill_prompt(
        load_prompt("lo_grouping"),
        Grade=grade,
        ChapterName=chapter,
        CCsAndLOs=ccs_and_los,
    )
    text = call_model(prompt, payload, max_output_tokens=24000)
    groups = source_exact_grouped_scripts(grouped_script_blocks(text), outcomes)
    return {
        "grouping": text,
        "groupingValidation": "",
        "groupingStatus": "Grouped",
        "groups": groups,
        "promptUsed": "lo_grouping",
        "validationPromptUsed": "",
    }


def generate_result(payload):
    outcomes_text = payload.get("learningOutcomes", "").strip()
    if not outcomes_text:
        raise ValueError("Missing grouped Learning Outcomes.")
    prompt = script_prompt(
        payload.get("grade", ""),
        payload.get("chapter", ""),
        payload.get("subtopic", ""),
        outcomes_text,
    )
    script = call_model(prompt, payload, max_output_tokens=8000)
    return {"script": script, "attempt": 1}


def validate_result(payload):
    script = payload.get("script", "").strip()
    outcomes_text = payload.get("learningOutcomes", "").strip()
    if not script or not outcomes_text:
        raise ValueError("Missing script or Learning Outcomes for validation.")
    prompt = validation_prompt(
        payload.get("grade", ""),
        payload.get("chapter", ""),
        payload.get("subtopic", ""),
        outcomes_text,
        script,
        payload.get("textbookReference", ""),
    )
    report = call_model(prompt, payload, max_output_tokens=8000)
    return {"report": report, "status": extract_verdict(report)}


def revise_result(payload):
    attempt = int(payload.get("attempt", 1))
    if attempt > MAX_REGENERATIONS:
        raise ValueError("Maximum regeneration limit reached.")

    outcomes_text = payload.get("learningOutcomes", "").strip()
    previous_script = payload.get("script", "").strip()
    report = payload.get("validationReport", "").strip()
    if not outcomes_text or not previous_script or not report:
        raise ValueError("Missing script, Learning Outcomes, or validation report.")

    prompt = (
        script_prompt(
            payload.get("grade", ""),
            payload.get("chapter", ""),
            payload.get("subtopic", ""),
            outcomes_text,
        )
        + "\n\nThe previous script was not approved. Regenerate a corrected version using the validation report below.\n"
        + "Do not preserve rejected lines. Produce only a fresh final script that can pass validation.\n\n"
        + f"Previous Script:\n{previous_script}\n\nValidation Report:\n{report}\n\n"
        + f"This is regeneration attempt {attempt} of {MAX_REGENERATIONS}."
    )
    script = call_model(prompt, payload, max_output_tokens=8000)
    return {"script": script, "attempt": attempt}


def run_approved_result(payload):
    outcomes_text = payload.get("learningOutcomes", "").strip()
    if not outcomes_text:
        raise ValueError("Missing grouped Learning Outcomes.")

    grade = payload.get("grade", "")
    chapter = payload.get("chapter", "")
    subtopic = payload.get("subtopic", "")
    textbook_reference = payload.get("textbookReference", "")

    history = []
    script = ""
    report = ""
    status = "Not Started"

    for attempt in range(1, MAX_REGENERATIONS + 1):
        if attempt == 1:
            prompt = script_prompt(grade, chapter, subtopic, outcomes_text)
        else:
            prompt = (
                script_prompt(grade, chapter, subtopic, outcomes_text)
                + "\n\nRegenerate using the previous validation report. Produce a fresh approved script.\n\n"
                + f"Previous Script:\n{script}\n\nValidation Report:\n{report}\n\n"
                + f"This is regeneration attempt {attempt} of {MAX_REGENERATIONS}."
            )
        try:
            script = call_model(prompt, payload, max_output_tokens=8000)
            report = call_model(
                validation_prompt(grade, chapter, subtopic, outcomes_text, script, textbook_reference),
                payload,
                max_output_tokens=8000,
            )
        except Exception as exc:
            raise ValueError(str(exc)) from exc

        status = extract_verdict(report)
        history.append({"attempt": attempt, "status": status, "script": script, "validationReport": report})
        if is_phase2_ready_status(status):
            break

    return {
        "status": status,
        "approved": status == "Approved",
        "phase2Ready": is_phase2_ready_status(status),
        "attemptsUsed": len(history),
        "maxRegenerations": MAX_REGENERATIONS,
        "script": script,
        "validationReport": report,
        "history": history,
    }


def combined_script_text_from_runs(script_runs):
    return "\n\n---\n\n".join(
        f"SCRIPT {index}: {item.get('title') or f'Script {index}'}\n"
        f"Status: {item.get('status') or 'Queued'}\n"
        f"Attempts used: {int(item.get('attemptsUsed') or 0)} / {MAX_REGENERATIONS}\n\n"
        f"{item.get('script') or ''}"
        for index, item in enumerate(script_runs, start=1)
        if item.get("script")
    )


def combined_validation_text_from_runs(script_runs):
    return "\n\n---\n\n".join(
        f"{item.get('title') or f'Script {index}'}\n{item.get('validationReport') or item.get('status') or 'Queued'}"
        for index, item in enumerate(script_runs, start=1)
    )


def script_batch_result(payload, job_id=None):
    groups = payload.get("scripts") or payload.get("groups") or []
    if not groups:
        raise ValueError("No grouped scripts were supplied for batch generation.")

    provider = model_provider_from_payload(payload)
    base_payload = {
        "grade": payload.get("grade", ""),
        "chapter": payload.get("chapter", ""),
        "subtopic": payload.get("subtopic", ""),
        "textbookReference": payload.get("textbookReference", ""),
        "modelProvider": provider,
        "claudeModel": payload.get("claudeModel"),
        "gemmaModel": payload.get("gemmaModel"),
        "claude_api_key": payload.get("claude_api_key", ""),
        "gemma_api_key": payload.get("gemma_api_key", ""),
    }
    script_runs = []
    for index, group in enumerate(groups):
        script_runs.append(
            {
                "index": index,
                "title": group.get("title") or f"Script {index + 1}",
                "learningOutcomes": group.get("learningOutcomes") or "",
                "subtopic": group.get("subtopic") or base_payload["subtopic"],
                "status": "Queued",
                "attemptsUsed": 0,
                "script": "",
                "validationReport": "",
                "history": [],
            }
        )

    def publish(current_index=0, phase="queued"):
        if not job_id:
            return
        update_job(
            job_id,
            progress={
                "phase": phase,
                "currentIndex": current_index,
                "total": len(script_runs),
                "scripts": script_runs,
                "script": combined_script_text_from_runs(script_runs),
                "validation": combined_validation_text_from_runs(script_runs),
            },
        )

    publish(0, "queued")
    for index, item in enumerate(script_runs):
        if not item["learningOutcomes"]:
            item["status"] = "Failed"
            item["validationReport"] = "Missing grouped Learning Outcomes."
            publish(index, "failed")
            continue

        item_payload = {
            **base_payload,
            "outcomes": [],
            "learningOutcomes": item["learningOutcomes"],
            "chapter": base_payload["chapter"] or item["title"],
            "subtopic": item["subtopic"] or base_payload["subtopic"],
        }

        try:
            item["status"] = "Generating"
            publish(index, "generating")
            generated = generate_result(item_payload)
            item["script"] = generated.get("script", "")
            item["attemptsUsed"] = max(int(item.get("attemptsUsed") or 0), int(generated.get("attempt") or 1))
            item["status"] = "Generated"
            item["history"].append({"attempt": generated.get("attempt") or 1, "status": "Generated", "script": item["script"]})
            publish(index, "generated")

            item["status"] = "Validating"
            publish(index, "validating")
            validation = validate_result({**item_payload, "script": item["script"]})
            item["status"] = validation.get("status") or "Needs Review"
            item["validationReport"] = validation.get("report", "")
            item["history"].append(
                {
                    "attempt": item["attemptsUsed"],
                    "status": item["status"],
                    "script": item["script"],
                    "validationReport": item["validationReport"],
                }
            )
            publish(index, "validated")

            while item["status"] == "Rejected" and item["attemptsUsed"] < MAX_REGENERATIONS:
                next_attempt = int(item["attemptsUsed"] or 0) + 1
                item["status"] = "Regenerating"
                publish(index, "regenerating")
                revised = revise_result(
                    {
                        **item_payload,
                        "script": item["script"],
                        "validationReport": item["validationReport"],
                        "attempt": next_attempt,
                    }
                )
                item["script"] = revised.get("script", "")
                item["attemptsUsed"] = int(revised.get("attempt") or next_attempt)
                item["history"].append({"attempt": item["attemptsUsed"], "status": "Regenerated", "script": item["script"]})
                publish(index, "generated")

                item["status"] = "Validating"
                publish(index, "validating")
                validation = validate_result({**item_payload, "script": item["script"]})
                item["status"] = validation.get("status") or "Needs Review"
                item["validationReport"] = validation.get("report", "")
                item["history"].append(
                    {
                        "attempt": item["attemptsUsed"],
                        "status": item["status"],
                        "script": item["script"],
                        "validationReport": item["validationReport"],
                    }
                )
                publish(index, "validated")
        except Exception as exc:
            item["status"] = "Failed"
            item["validationReport"] = str(exc) or exc.__class__.__name__
            item["history"].append({"attempt": item["attemptsUsed"] or 1, "status": "Failed", "validationReport": item["validationReport"]})
            publish(index, "failed")
            continue

    approved = all(item.get("status") == "Approved" for item in script_runs)
    phase2_ready = all(is_phase2_ready_status(item.get("status")) for item in script_runs)
    result = {
        "status": "Approved" if approved else ("Ready for Phase 2" if phase2_ready else "Needs Review"),
        "approved": approved,
        "phase2Ready": phase2_ready,
        "scripts": script_runs,
        "script": combined_script_text_from_runs(script_runs),
        "validation": combined_validation_text_from_runs(script_runs),
        "history": [history_item for item in script_runs for history_item in item.get("history", [])],
        "total": len(script_runs),
    }
    publish(len(script_runs), "complete")
    return result


@app.post("/api/group")
def group_route():
    try:
        return jsonify(group_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/generate")
def generate_route():
    try:
        return jsonify(generate_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/validate")
def validate_route():
    try:
        return jsonify(validate_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/revise")
def revise_route():
    try:
        return jsonify(revise_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/run-approved")
def run_approved_route():
    try:
        return jsonify(run_approved_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


def animation_storyboard_result(payload):
    prompt = animation_prompt(payload, "storyboard")
    storyboard = call_model(prompt, payload, max_output_tokens=int(payload.get("maxTokens") or 16000))
    return {
        "storyboard": storyboard,
        "provider": model_provider_from_payload(payload),
        "model": model_name_from_payload(payload),
        "promptUsed": "animation_phase",
    }


def animation_code_result(payload):
    prompt = animation_prompt(payload, "code")
    raw_code = call_model(prompt, payload, max_output_tokens=int(payload.get("maxTokens") or 24000))
    code = extract_python_code(raw_code)
    saved_file = save_generated_python_script(payload.get("title") or "animation_scene", code)
    render_job = launch_start_renderer([saved_file])
    return {
        "code": code,
        "raw": raw_code,
        "savedFile": saved_file,
        "renderJob": render_job,
        "provider": model_provider_from_payload(payload),
        "model": model_name_from_payload(payload),
        "promptUsed": "animation_phase",
    }


def animation_package_result(payload):
    prompt = animation_prompt(payload, "package")
    package = call_model(prompt, payload, max_output_tokens=int(payload.get("maxTokens") or 30000))
    code = extract_python_code(package)
    saved_file = save_generated_python_script(payload.get("title") or "animation_scene", code)
    render_job = launch_start_renderer([saved_file])
    return {
        "package": package,
        "code": code,
        "savedFile": saved_file,
        "renderJob": render_job,
        "provider": model_provider_from_payload(payload),
        "model": model_name_from_payload(payload),
        "promptUsed": "animation_phase",
    }


def combined_storyboard_text_from_animation_runs(animation_runs):
    return "\n\n---\n\n".join(
        f"## STORYBOARD {index}: {item.get('title') or f'Script {index}'}\n\n{item.get('storyboard') or ''}"
        for index, item in enumerate(animation_runs, start=1)
        if item.get("storyboard")
    )


def combined_python_text_from_animation_runs(animation_runs):
    return "\n\n# ---\n\n".join(
        f"# PYTHON {index}: {item.get('title') or f'Script {index}'}\n\n{item.get('code') or ''}"
        for index, item in enumerate(animation_runs, start=1)
        if item.get("code")
    )


def compact_animation_runs(animation_runs):
    compacted = []
    for item in animation_runs:
        output = {
            "index": item.get("index"),
            "title": item.get("title"),
            "transcript": item.get("transcript") or "",
            "storyboardStatus": item.get("storyboardStatus") or "Storyboard Queued",
            "codeStatus": item.get("codeStatus") or "Code Not Started",
        }
        if item.get("storyboard"):
            output["storyboard"] = item["storyboard"]
        if item.get("code"):
            output["code"] = item["code"]
        if item.get("savedFile"):
            output["savedFile"] = item["savedFile"]
        if item.get("error"):
            output["error"] = item["error"]
        compacted.append(output)
    return compacted


def animation_batch_result(payload, job_id=None):
    raw_items = payload.get("items") or []
    if not raw_items:
        raise ValueError("No Phase 2 scripts were supplied for animation generation.")

    mode = str(payload.get("mode") or "code").strip().lower()
    if mode not in {"storyboard", "code"}:
        raise ValueError("Animation batch mode must be storyboard or code.")

    provider = model_provider_from_payload(payload)
    base_payload = {
        "subject": payload.get("subject") or "Mathematics",
        "grade": payload.get("grade") or "",
        "duration": payload.get("duration") or "90 seconds",
        "contentType": payload.get("contentType") or "Concept Video",
        "modelProvider": provider,
        "claudeModel": payload.get("claudeModel"),
        "gemmaModel": payload.get("gemmaModel"),
        "claude_api_key": payload.get("claude_api_key", ""),
        "gemma_api_key": payload.get("gemma_api_key", ""),
    }
    animation_runs = []
    for index, item in enumerate(raw_items):
        storyboard = str(item.get("storyboard") or "").strip()
        code = str(item.get("code") or "").strip()
        storyboard_status = item.get("storyboardStatus") or ("Storyboard Done" if storyboard else "Storyboard Queued")
        code_status = item.get("codeStatus") or ("Code Done" if code else ("Code Queued" if mode == "code" else "Code Not Started"))
        if storyboard_status in {"Failed", "Paused", "Storyboard Failed", "Storyboard Paused"} and storyboard:
            storyboard_status = "Storyboard Done"
        if code_status in {"Failed", "Paused", "Code Failed", "Code Paused"} and code:
            code_status = "Code Done"
        animation_runs.append(
            {
                "index": item.get("index", index),
                "title": item.get("title") or f"Script {index + 1}",
                "transcript": item.get("transcript") or "",
                "storyboard": storyboard,
                "code": code,
                "savedFile": item.get("savedFile"),
                "storyboardStatus": storyboard_status,
                "codeStatus": code_status,
                "error": item.get("error") or "",
            }
        )

    disk_used_names = set()

    def saved_files():
        return [item["savedFile"] for item in animation_runs if item.get("savedFile")]

    def current_result(status="Running", phase="queued", current_index=0, render_job=None):
        return {
            "status": status,
            "mode": mode,
            "phase": phase,
            "currentIndex": current_index,
            "total": len(animation_runs),
            "items": compact_animation_runs(animation_runs),
            "storyboard": combined_storyboard_text_from_animation_runs(animation_runs),
            "pythonCode": combined_python_text_from_animation_runs(animation_runs),
            "savedFiles": saved_files(),
            "renderJob": render_job,
            "provider": provider,
            "model": model_name_from_payload(payload),
            "promptUsed": "animation_phase",
        }

    def publish(current_index=0, phase="queued", status="Running"):
        if not job_id:
            return
        update_job(job_id, progress=current_result(status=status, phase=phase, current_index=current_index))

    publish(0, "queued")

    for index, item in enumerate(animation_runs):
        if item.get("storyboard") and item.get("storyboardStatus") == "Storyboard Done":
            publish(index, "storyboard-skipped")
            continue
        if not str(item.get("transcript") or "").strip():
            item["storyboardStatus"] = "Storyboard Failed"
            item["error"] = "Missing transcript for this script."
            publish(index, "storyboard-failed", "Paused")
            return current_result(status="Paused", phase="storyboard-failed", current_index=index)
        try:
            item["storyboardStatus"] = "Storyboard Generating"
            item["codeStatus"] = item.get("codeStatus") or ("Code Queued" if mode == "code" else "Code Not Started")
            item["error"] = ""
            publish(index, "storyboard", "Running")
            storyboard_payload = {
                **base_payload,
                "title": item["title"],
                "transcript": item["transcript"],
            }
            storyboard = call_model(
                animation_prompt(storyboard_payload, "storyboard"),
                storyboard_payload,
                max_output_tokens=int(payload.get("storyboardMaxTokens") or payload.get("maxTokens") or 16000),
            )
            item["storyboard"] = storyboard
            item["storyboardStatus"] = "Storyboard Done"
            publish(index, "storyboard-done", "Running")
        except Exception as exc:
            item["storyboardStatus"] = "Storyboard Failed"
            item["error"] = str(exc) or exc.__class__.__name__
            publish(index, "storyboard-failed", "Paused")
            return current_result(status="Paused", phase="storyboard-failed", current_index=index)

    if mode == "storyboard":
        publish(len(animation_runs), "complete", "Generated")
        return current_result(status="Generated", phase="complete", current_index=len(animation_runs))

    publish(0, "code-queued", "Running")
    for index, item in enumerate(animation_runs):
        if item.get("code") and item.get("codeStatus") == "Code Done":
            publish(index, "code-skipped")
            continue
        try:
            item["codeStatus"] = "Code Generating"
            item["error"] = ""
            publish(index, "code", "Running")
            code_payload = {
                **base_payload,
                "title": item["title"],
                "transcript": item["transcript"],
                "storyboard": item.get("storyboard") or "",
            }
            raw_code = call_model(
                animation_prompt(code_payload, "code"),
                code_payload,
                max_output_tokens=int(payload.get("codeMaxTokens") or payload.get("maxTokens") or 24000),
            )
            code = extract_python_code(raw_code)
            saved_file = save_generated_python_script(item["title"], code, disk_used_names)
            item["code"] = code
            item["savedFile"] = saved_file
            item["codeStatus"] = "Code Done"
            publish(index, "code-done", "Running")
        except Exception as exc:
            item["codeStatus"] = "Code Failed"
            item["error"] = str(exc) or exc.__class__.__name__
            publish(index, "code-failed", "Paused")
            return current_result(status="Paused", phase="code-failed", current_index=index)

    render_job = launch_start_renderer(saved_files())
    publish(len(animation_runs), "complete", "Generated")
    return current_result(status="Generated", phase="complete", current_index=len(animation_runs), render_job=render_job)


@app.post("/api/animation/storyboard")
def animation_storyboard_route():
    try:
        return jsonify(animation_storyboard_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/animation/code")
def animation_code_route():
    try:
        return jsonify(animation_code_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/animation/package")
def animation_package_route():
    try:
        return jsonify(animation_package_result(payload_with_request_key(request.get_json(force=True))))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


JOB_HANDLERS = {
    "group": group_result,
    "generate": generate_result,
    "validate": validate_result,
    "revise": revise_result,
    "run-approved": run_approved_result,
    "script-batch": script_batch_result,
    "animation-storyboard": animation_storyboard_result,
    "animation-code": animation_code_result,
    "animation-package": animation_package_result,
    "animation-batch": animation_batch_result,
}


def compact_job(job):
    output = {
        "job_id": job["job_id"],
        "kind": job["kind"],
        "status": job["status"],
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
    }
    if job["status"] == "done":
        output["result"] = job.get("result") or {}
    if job.get("progress") is not None:
        output["progress"] = job.get("progress")
    if job["status"] == "failed":
        output["error"] = job.get("error") or "Job failed."
    return output


def cleanup_old_jobs():
    cutoff = time.time() - JOB_TTL_SECONDS
    with JOB_LOCK:
        for job_id in [job_id for job_id, job in JOBS.items() if job.get("updated_at", 0) < cutoff]:
            JOBS.pop(job_id, None)


def update_job(job_id, **changes):
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return
        job.update(changes)
        job["updated_at"] = time.time()


def run_job(job_id, kind, payload):
    update_job(job_id, status="running")
    try:
        if kind == "script-batch":
            result = script_batch_result(payload, job_id=job_id)
        elif kind == "animation-batch":
            result = animation_batch_result(payload, job_id=job_id)
        else:
            result = JOB_HANDLERS[kind](payload)
        update_job(job_id, status="done", result=result)
    except Exception as exc:
        update_job(job_id, status="failed", error=str(exc) or exc.__class__.__name__)


@app.post("/api/jobs/<kind>")
def start_job_route(kind):
    if kind not in JOB_HANDLERS:
        return jsonify({"error": "Unknown job type."}), 404
    try:
        payload = payload_with_request_key(request.get_json(force=True))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    cleanup_old_jobs()
    job_id = uuid.uuid4().hex
    job = {
        "job_id": job_id,
        "kind": kind,
        "status": "queued",
        "created_at": time.time(),
        "updated_at": time.time(),
    }
    with JOB_LOCK:
        JOBS[job_id] = job
    JOB_EXECUTOR.submit(run_job, job_id, kind, payload)
    return jsonify(compact_job(job)), 202


@app.get("/api/jobs/<job_id>")
def job_status_route(job_id):
    cleanup_old_jobs()
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"error": "Job not found or expired."}), 404
        return jsonify(compact_job(job))


@app.post("/api/bulk/python-scripts")
def bulk_python_scripts_route():
    payload = payload_with_request_key(request.get_json(force=True))
    outcomes = payload.get("outcomes") or []
    if not outcomes:
        return jsonify({"error": "Select at least one Learning Outcome for bulk generation."}), 400

    try:
        if model_provider_from_payload(payload) == "gemma":
            gemma_api_key_from_payload(payload)
        else:
            claude_api_key_from_payload(payload)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    duration = payload.get("duration") or "90 seconds"
    content_type = payload.get("contentType") or "Concept Video"
    groups = bulk_groups_from_outcomes(outcomes)
    max_items = int(payload.get("maxItems") or len(groups))
    selected_groups = groups[:max_items]

    used_names = set()
    disk_used_names = set()
    saved_files = []
    manifest = []
    archive_buffer = BytesIO()

    with zipfile.ZipFile(archive_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, group in enumerate(selected_groups, start=1):
            grade = str(group.get("grade") or payload.get("grade") or "").strip()
            chapter = str(group.get("chapter") or payload.get("chapter") or "").strip()
            subtopic = str(group.get("subtopic") or payload.get("subtopic") or "").strip()
            title = script_animation_title(index, group)
            outcomes_text = bulk_learning_outcomes_text(group["outcomes"])
            filename = unique_python_filename(title, used_names)
            stem = filename[:-3]
            if not outcomes_text:
                manifest.append({"index": index, "status": "skipped", "title": title, "error": "Missing Learning Outcome text."})
                continue

            try:
                grouped_los = call_model(
                    fill_prompt(
                        load_prompt("lo_grouping"),
                        Grade=grade,
                        ChapterName=chapter,
                        CCsAndLOs=format_ccs_and_los(group["outcomes"]),
                    ),
                    payload,
                    max_output_tokens=8000,
                )

                transcript = ""
                validation_report = ""
                status = "Not Started"
                for attempt in range(1, MAX_REGENERATIONS + 1):
                    if attempt == 1:
                        transcript_prompt = script_prompt(grade, chapter, subtopic, grouped_los)
                    else:
                        transcript_prompt = (
                            script_prompt(grade, chapter, subtopic, grouped_los)
                            + "\n\nThe previous transcript was not approved. Regenerate using this validation report.\n\n"
                            + f"Previous Transcript:\n{transcript}\n\nValidation Report:\n{validation_report}\n\n"
                            + f"This is regeneration attempt {attempt} of {MAX_REGENERATIONS}."
                        )
                    transcript = call_model(transcript_prompt, payload, max_output_tokens=8000)
                    validation_report = call_model(
                        validation_prompt(grade, chapter, subtopic, grouped_los, transcript, payload.get("textbookReference", "")),
                        payload,
                        max_output_tokens=8000,
                    )
                    status = extract_verdict(validation_report)
                    if is_phase2_ready_status(status):
                        break

                archive.writestr(f"{stem}_grouped_los.txt", grouped_los)
                archive.writestr(f"{stem}_validated_transcript.txt", transcript)
                archive.writestr(f"{stem}_validation_report.txt", validation_report)

                if not is_phase2_ready_status(status):
                    manifest.append(
                        {
                            "index": index,
                            "status": "validation_failed",
                            "validation_status": status,
                            "title": title,
                            "rows": group.get("rows"),
                            "transcript": f"{stem}_validated_transcript.txt",
                            "validation_report": f"{stem}_validation_report.txt",
                        }
                    )
                    continue

                storyboard_payload = {
                    **payload,
                    "transcript": transcript,
                    "title": title,
                    "subject": group.get("subject") or "Mathematics",
                    "grade": grade,
                    "duration": duration,
                    "contentType": content_type,
                }
                storyboard = call_model(
                    animation_prompt(storyboard_payload, "storyboard"),
                    storyboard_payload,
                    max_output_tokens=16000,
                )
                code_payload = {**storyboard_payload, "storyboard": storyboard}
                code_response = call_model(
                    animation_prompt(code_payload, "code"),
                    code_payload,
                    max_output_tokens=24000,
                )
                archive.writestr(f"{stem}_storyboard.txt", storyboard)
                code = extract_python_code(code_response)
                saved_file = save_generated_python_script(title, code, disk_used_names)
                saved_files.append(saved_file)
                archive.writestr(filename, code + "\n")
                manifest.append(
                    {
                        "index": index,
                        "status": "generated",
                        "validation_status": status,
                        "filename": filename,
                        "saved_path": saved_file["path"],
                        "title": title,
                        "grade": grade,
                        "chapter": chapter,
                        "subtopic": subtopic,
                        "rows": group.get("rows"),
                        "grouped_los": f"{stem}_grouped_los.txt",
                        "validated_transcript": f"{stem}_validated_transcript.txt",
                        "validation_report": f"{stem}_validation_report.txt",
                        "storyboard": f"{stem}_storyboard.txt",
                    }
                )
            except Exception as exc:
                manifest.append(
                    {
                        "index": index,
                        "status": "failed",
                        "title": title,
                        "rows": group.get("rows"),
                        "error": str(exc),
                    }
                )

        render_job = launch_start_renderer(saved_files)
        archive.writestr("manifest.json", json.dumps({"items": manifest, "render_job": render_job}, ensure_ascii=False, indent=2))

    archive_buffer.seek(0)
    return Response(
        archive_buffer.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": "attachment; filename=bulk_python_scripts.zip"},
    )


@app.post("/api/download")
def download_route():
    payload = request.get_json(force=True)
    scripts = payload.get("scripts") or []
    ready = [script for script in scripts if is_phase2_ready_status(script.get("status"))]
    if len(ready) != len(scripts):
        return jsonify({"error": "Download is blocked until every selected script is Approved or Needs Minor Revision."}), 400
    body = json.dumps({"phase2_ready_scripts": ready}, ensure_ascii=False, indent=2)
    return Response(
        body,
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=approved_conceptual_scripts.json"},
    )
